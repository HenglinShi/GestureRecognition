'''
Created on Dec 15, 2016

@author: hshi
'''
import os
import numpy as np
import caffe
import shutil
from DataPreparation import myMkdir, moveFile, extractFeature
from DefineNet import defineNet
from DefineSolver import defineSolver
from sklearn import preprocessing
import cPickle as pickle
from CrossValidatingRandomSampler import CrossValidatingRandomSampler
from TransitionData import TransitionData
from Train import train, train_earlyStop
from Tuples import Tuples
from Test import testOneNet
import openpyxl as pyxl
from Trainging_Lib import actionUnits

def report(sheet, currentResult):
            
    sheet['D1'] = "Best Valid Accuracy"
    sheet['D2'] = currentResult.getBestValidAccuracy()
            
    sheet['E1'] = "bestWeightInd"
    sheet['E2'] = currentResult.getBestWeightInd()
            
    sheet['F1'] = "Best Valid Weight"
    sheet['F2'] = currentResult.getBestWeightPath()
            
            
    sheet['G1'] = "Train Ratio"
    sheet['G2'] = currentResult.getRatio_train()
            
    sheet['H1'] = "Valid Ratio"
    sheet['H2'] = currentResult.getRatio_valid()
    
    sheet['I1'] = "folds num"
    sheet['I2'] = currentResult.getFoldsNum()
            
    sheet['J1'] = "Testing Accuracy of Best Valid Weight"
    sheet['J2'] = currentResult.getBestWeightTestAccuracy()
            
    sheet['K1'] = "Batch Size"
    sheet['K2'] = currentResult.getBatchSize()
            
    sheet['L1'] = "Training Epoches"
    sheet['L2'] = currentResult.getTrainingEpoches()
            
    sheet['M1'] = "Normalization Type"
    sheet['M2'] = currentResult.getNormalizationType()
            
    sheet['N1'] = "sample prenormalization"
    sheet['N2'] = currentResult.getSampleWisePreNormalization()
    

    sheet['O1'] = "minvalid Loss"

    validAccuracy = currentResult.getValidAccuracy()
    weightFilePath = currentResult.getWeightPath()
    testAccuracy = currentResult.getTestAccuracy() 
    validLoss = currentResult.getValidLoss()
            
    sheet['A1'] = "Validating Accuracy"
    sheet['B1'] = "Weight File Path"
    sheet['C1'] = "Testing Accuracy"
            
    for row in range(len(weightFilePath)):
        _ = sheet.cell(column = 1, row = row + 2, value = validAccuracy[row])
        #sheet['B'+str(row+2)] = weightFilePath[row]
        _ = sheet.cell(column = 2, row = row + 2, value = weightFilePath[row])
        #_ = sheet.cell(column = 2, row = row + 2, value = "%s" % weightFilePath[row])
        _ = sheet.cell(column = 3, row = row + 2, value = testAccuracy[row])
        
        _ = sheet.cell(column = 15, row = row + 2, value = validLoss[row])



    return sheet

def summarizeReport(sheet, result):
    sheet['A1'] = "fold_no"
    sheet['B1'] = "Best Valid Accuracy"
    sheet['C1'] = "Testing Accuracy of Best Valid Weight"
    sheet['D1'] = "bestWeightInd"        
    sheet['E1'] = "Best Valid Weight"
    
    sheet['F1'] = "Train Ratio"
    sheet['G1'] = "Valid Ratio"
    sheet['H1'] = "Batch Size"            
    sheet['I1'] = "Training Epoches"     
    sheet['J1'] = "Normalization Type"
    
    sheet['K1'] = "sample prenormalize"
    sheet['L1'] = "stateNumPERclass"
    sheet['M1'] = "bestValidLoss"
    
  

    tmpSampleNames_test = list()
    tmpPaths = list()
    
    for row in range(len(result)):
        currentResult = result[row]

                
        _ = sheet.cell(column = 1, row = row + 2, value = row + 1)
        _ = sheet.cell(column = 2, row = row + 2, value = currentResult.getBestValidAccuracy())
        _ = sheet.cell(column = 3, row = row + 2, value = currentResult.getBestWeightTestAccuracy())
        _ = sheet.cell(column = 4, row = row + 2, value = currentResult.getBestWeightInd())
        _ = sheet.cell(column = 5, row = row + 2, value = currentResult.getBestWeightPath()) 
           
        _ = sheet.cell(column = 6, row = row + 2, value = currentResult.getRatio_train())   
        _ = sheet.cell(column = 7, row = row + 2, value = currentResult.getRatio_valid())   
  
        
        _ = sheet.cell(column = 8, row = row + 2, value = currentResult.getBatchSize())          
        _ = sheet.cell(column = 9, row = row + 2, value = currentResult.getTrainingEpoches())     
        _ = sheet.cell(column = 10, row = row + 2, value = currentResult.getNormalizationType())     
        _ = sheet.cell(column = 11, row = row + 2, value = currentResult.getSampleWisePreNormalization())    
         
        _ = sheet.cell(column = 12, row = row + 2, value = currentResult.getStateTypeNumPerClass()) 
        
        
        sampleNames_test = currentResult.getSampleNames_test()
        paths = currentResult.getHMMpaths()
        
        if row == 0:
            tmpLabels = currentResult.getLabel_test()
            
        else:
            tmpLabels = np.concatenate((tmpLabels, currentResult.getLabel_test()), axis = 0)
            
        for i in range(len(sampleNames_test)):
            tmpSampleNames_test.append(sampleNames_test[i])        
            tmpPaths.append(paths[i])
            
            
        
    resultNum = len(result)
    
    for i in range(len(tmpSampleNames_test)):     
            
        _ = sheet.cell(column = 1, row = resultNum + 5 + i, value = tmpSampleNames_test[i])     
        _ = sheet.cell(column = 2, row = resultNum + 5 + i, value = tmpLabels[i])   
        
        currentPath = tmpPaths[i]
        
        for j in range(currentPath.shape[0]):  
            _ = sheet.cell(column = 3 + j, row = resultNum + 5 + i, value = currentPath[j])   
    



    return sheet


def main():
    
    caffe.set_device(2)
    caffe.set_mode_gpu()
    
    actionUnits = ['ElbowLeft', 'WristLeft', 'ShoulderLeft','HandLeft',
                'ElbowRight', 'WristRight','ShoulderRight','HandRight',
                    'Head','Spine','HipCenter']
    
    earlyStop = 0
    ratio_train = 0.65
    ratio_valid = 0.15
    n_folds = 5
    stateTypeNumPerClass = 10
    classTypeNum = 20
    stateTypeNumAll = stateTypeNumPerClass * classTypeNum
    transitionEstimationType = 2
    normalizationType = 2
    sampleWisePreNormalization = 0
    subjectWisePreNormalization = 0
    batchSize = 400
    trainingEpoches = 150
    
    
    
    rootDir = "/research/tklab/personal/hshi/Gesture Recognition/Experiments/"
    workingDirName = "transition(" + str(transitionEstimationType) + ")_" + \
                     "normalize(" + str(normalizationType) + ")_" + \
                     "batchSize(" + str(batchSize) + ")_" + \
                     "trainingEpoches(" + str(trainingEpoches) + ")_" + \
                     "states(" + str(stateTypeNumPerClass) + ")_" + \
                     "nflods(" + str(n_folds) + ")_" + \
                     "earlyStop_1218"
                     
    workingDir = os.path.join(rootDir, workingDirName)
    rawDataDir = "/research/tklab/personal/hshi/Gesture Recognition/chalearn/"
    myMkdir(workingDir)
    
    lossOriented = 0
    
    

    
    
    samplingResults = list()
    #sampleDir = os.path.join(rawDataDir, 'data')
    sampleDir = rawDataDir
    
    sampleNames = os.listdir(sampleDir)
    mSampler = CrossValidatingRandomSampler(sampleNames_all=sampleNames,
                                            ratio_train=ratio_train,
                                            ratio_valid=ratio_valid,
                                            n_folds=n_folds)

    currentSamplingResult = mSampler.getNextSamplingResult()
    
    
    result = list()
    while currentSamplingResult != None:
        currentResult = Tuples()
        
        samplingResults.append(samplingResults)
        sampleNames_train = currentSamplingResult.getSampleNames_train()
        sampleNames_valid = currentSamplingResult.getSampleNames_valid()
        sampleNames_test = currentSamplingResult.getSampleNames_test()
        

        
        currentWorkingDir = os.path.join(workingDir, str(currentSamplingResult.getCVite()))
        myMkdir(currentWorkingDir)
        
        currentWeightDir = os.path.join(currentWorkingDir, 'weight')
        currentWorkingDataDir = os.path.join(currentWorkingDir, 'data')
        currentWorkingDataDir_train = os.path.join(currentWorkingDataDir, 'train')
        currentWorkingDataDir_valid = os.path.join(currentWorkingDataDir, 'valid')
        currentWorkingDataDir_test = os.path.join(currentWorkingDataDir, 'test')
        
        myMkdir(currentWeightDir)
        myMkdir(currentWorkingDataDir)
        myMkdir(currentWorkingDataDir_train)
        myMkdir(currentWorkingDataDir_valid)
        myMkdir(currentWorkingDataDir_test)
        
        moveFile(sampleNames_train, sampleDir, currentWorkingDataDir_train)
        moveFile(sampleNames_valid, sampleDir, currentWorkingDataDir_valid)
        moveFile(sampleNames_test, sampleDir, currentWorkingDataDir_test)
        

        
        # extract feature
        sample_train, label_train, clipMarker_train = extractFeature(currentWorkingDataDir_train, actionUnits, classTypeNum, stateTypeNumPerClass, sampleWisePreNormalization)  
        sample_valid, label_valid, clipMarker_valid = extractFeature(currentWorkingDataDir_valid, actionUnits, classTypeNum, stateTypeNumPerClass, sampleWisePreNormalization) 
        sample_test, label_test, clipMarker_test = extractFeature(currentWorkingDataDir_test, actionUnits, classTypeNum, stateTypeNumPerClass, sampleWisePreNormalization)
        

        # Transition parameter
        if transitionEstimationType == 2:
            # using training and validating data to estimate the prior
            currentTransitionDataDir = os.path.join(currentWorkingDataDir, 'transitionData')
            myMkdir(currentTransitionDataDir)
            moveFile(sampleNames_train, currentWorkingDataDir_train, currentTransitionDataDir)
            moveFile(sampleNames_valid, currentWorkingDataDir_valid, currentTransitionDataDir)
            
        if transitionEstimationType == 1:
            # just using trianing data to estimate the prior
            currentTransitionDataDir = currentWorkingDataDir_valid
            
        transitionData = TransitionData(currentTransitionDataDir,
                                        "transitionData",
                                        classTypeNum,
                                        stateTypeNumPerClass)
            
        

        if subjectWisePreNormalization:
            pass
        else:
            pass
        
        scalar = None
        if normalizationType == 2:
            sample_train_valid = np.concatenate((sample_train,
                                                 sample_valid),
                                                axis = 0)
            
            scalar = preprocessing.StandardScaler().fit(sample_train_valid)

        if normalizationType == 1:
            scalar = preprocessing.StandardScaler().fit(sample_train)

        mean = scalar.mean_
        std = scalar.scale_
        
        sample_train = scalar.transform(sample_train)
        sample_valid = scalar.transform(sample_valid)
        sample_test = scalar.transform(sample_test)
        
        
        
        dataPath_train = os.path.join(currentWorkingDataDir, 'train.pkl')
        dataPath_valid = os.path.join(currentWorkingDataDir, 'valid.pkl')
        dataPath_test = os.path.join(currentWorkingDataDir, 'test.pkl')
        #currentTransitionPath = transitionData.getPath()
        
        f = open(dataPath_train,'wb')
        pickle.dump( {"sample": sample_train, "label": label_train, "clipMarker": clipMarker_train},f) 
        f.close() 
        
        
        f = open(dataPath_valid,'wb')
        pickle.dump( {"sample": sample_valid, "label": label_valid,  "clipMarker": clipMarker_valid},f)
        f.close() 
        
        f = open(dataPath_test,'wb')
        pickle.dump( {"sample": sample_test, "label": label_test, "clipMarker": clipMarker_test},f)
        f.close()      
        
        currentDistributionPath = os.path.join(currentWorkingDir, 'distribution.pkl')
        f = open(currentDistributionPath,'wb')
        pickle.dump( {"Mean": mean, "Std": std },f)
        f.close() 
        
        
        netPath_train = os.path.join(currentWorkingDir, "net_train.prototxt")
        netPath_valid = os.path.join(currentWorkingDir, "net_valid.prototxt")
        netPath_test = os.path.join(currentWorkingDir, "net_test.prototxt")
        solverPath = os.path.join(currentWorkingDir, "solver.prototxt")
        
        dataFileParam_train = dict(data_path = dataPath_train, batch_size = batchSize)
        dataFileParam_valid = dict(data_path = dataPath_valid, batch_size = batchSize)
        dataFileParam_test = dict(data_path = dataPath_test, batch_size = batchSize)
    
        net_train = defineNet('train', batchSize, dataFileParam_train, stateTypeNumAll, 528)
        net_valid = defineNet('test', batchSize, dataFileParam_valid, stateTypeNumAll, 528)
        net_test = defineNet('test', batchSize, dataFileParam_test, stateTypeNumAll, 528)
        

        solver = defineSolver(netPath_train, netPath_valid, 0.01, './')
        
        with open(netPath_train, 'w') as f:
            f.write(str(net_train.to_proto()))
          
        with open(netPath_valid, 'w') as f:   
            f.write(str(net_valid.to_proto()))
                
        with open(netPath_test, 'w') as f:   
            f.write(str(net_test.to_proto()))
        
        with open(solverPath, 'w') as f:
            f.write(str(solver))
            
        data_valid = {'sample': sample_valid,
                      'label': label_valid,
                      'clipMarker': clipMarker_valid}
        
        
        # Train
        #weightDir = os.path.join(workingDir, 'weight')
        
        
        
        # Test
        #
        if earlyStop:
            
            validAccuracy, weightPath, validLoss = train_earlyStop(currentWeightDir, 
                                          solverPath, 
                                          netPath_train, 
                                          netPath_valid, 
                                          batchSize, 
                                          sample_train.shape[0], 
                                          sample_valid.shape[0], 
                                          trainingEpoches, 
                                          data_valid)
            

            currentBestWeightInd = len(weightPath) - 1
            currentBestValidAccuracy = validAccuracy[currentBestWeightInd]
            
        else:
            
            validAccuracy, weightPath, validLoss = train(currentWeightDir, 
                                          solverPath, 
                                          netPath_train, 
                                          netPath_valid, 
                                          batchSize, 
                                          sample_train.shape[0], 
                                          sample_valid.shape[0], 
                                          trainingEpoches, 
                                          data_valid)
            
            if lossOriented:
                currentBestWeightInd = validLoss.argmin()
                currentBestValidAccuracy = validLoss[currentBestWeightInd]
            else:
                currentBestWeightInd = validAccuracy.argmax()   
                currentBestValidAccuracy = validAccuracy[currentBestWeightInd]
               
        
        
        currentBestWeightPath = weightPath[currentBestWeightInd]
        
        currentWeightPredictions, \
        currentLabels, \
        currentPredictedPaths,\
        tmpSampleNames_test = testOneNet(netPath_test, 
                                           currentBestWeightPath, 
                                           transitionData, 
                                           scalar, 
                                           currentWorkingDataDir_test, 
                                           stateTypeNumAll, 
                                           stateTypeNumPerClass, 
                                           batchSize, 
                                           sampleWisePreNormalization, 
                                           subjectWisePreNormalization, 
                                           transitionEstimationType, 
                                           normalizationType, 
                                           actionUnits)
        
        
        currentTestAccuracy = np.zeros(shape = validAccuracy.shape, dtype = np.float)
        
        
        bestValidTestAccuracy = (float) (sum(currentWeightPredictions == currentLabels)) / len(currentLabels)
        currentTestAccuracy[currentBestWeightInd] = bestValidTestAccuracy
        
        
        
        
        currentResult.setValidLoss(validLoss)
        currentResult.setLabel_test(currentLabels)
        currentResult.setSampleNames_test(tmpSampleNames_test)
        currentResult.setValidAccuracy(validAccuracy)
        currentResult.setWeightPath(weightPath)
        currentResult.setSampleNames_train(sampleNames_train)
        currentResult.setSampleNames_valid(sampleNames_valid)
        currentResult.setSampleNames_test(sampleNames_test)
        currentResult.setWorkingDir(currentWorkingDataDir)
        currentResult.setTransitionEstimationType(transitionEstimationType)
        currentResult.setNormalizationType(normalizationType)
        currentResult.setSampleWisePreNormalization(sampleWisePreNormalization)
        currentResult.setSubjectWisePreNormalization(subjectWisePreNormalization)
        currentResult.setTrainingEpoches(trainingEpoches)
        currentResult.setBatchSize(batchSize)
        
        currentResult.setStateTypeNumPerClass(stateTypeNumPerClass)
        currentResult.setBestWeightInd(currentBestWeightInd)
        currentResult.setBestValidAccuracy(currentBestValidAccuracy)
        currentResult.setBestWeightPath(currentBestWeightPath)
        currentResult.setHMMpaths(currentPredictedPaths)
        currentResult.setTestAccuracy(currentTestAccuracy)
        currentResult.setBestWeightTestAccuracy(bestValidTestAccuracy)
        currentResult.setRatio_train(ratio_train)
        currentResult.setRatio_valid(ratio_valid)
        currentResult.setFoldsNum(n_folds)
        
        result.append(currentResult)
        currentSamplingResult = mSampler.getNextSamplingResult()
        
        
        shutil.rmtree(currentWorkingDataDir)
        
        
        
        
        
        

    resultFilePath = os.path.join(rootDir, workingDirName + 'result.xlsx')
        
    resultFile = pyxl.Workbook()
    sheet1 = resultFile.active
    sheet1.title = 'sum'
 
    for i in range(len(result)):   
        currentResult = result[i]
        newsheet = resultFile.create_sheet(str(i + 1), i + 2) 
        newsheet = report(newsheet, currentResult)
    
    sheet1 = summarizeReport(sheet1, result)

            
    resultFile.save(resultFilePath)
    
    
    print (12)
        
        
        
        
        
        
        
        
        
        
        
        
        
        



if __name__ == '__main__':
    main()